"""
Genera la planilla para que Gianluca/Fabricio carguen el conteo físico de
stock a mano. Parte del catálogo actual (stock_con_precios.xlsx) como
punto de partida — así no tienen que retipear ~290 productos — pero deja
la columna de stock en blanco a propósito: el objetivo es un conteo real,
no confirmar los números viejos (que ya sabemos desactualizados).

El archivo resultante usa exactamente las columnas que espera el
importador ya construido en tq-ecommerce
(scripts/import-inventory-lib.ts::parseRows): nombre, codigo, categoria,
talla, stock, precio_venta, precio_costo. No hace falta ninguna conversión
adicional antes de subirlo. La columna 'codigo' es opcional: si se
completa con el código del proveedor, el producto queda mapeado para que
las próximas compras (Cuadro de Pedido) lo reconozcan automáticamente.

Uso: python generar_plantilla_stock.py
Salida: plantilla_carga_stock.xlsx
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ORIGEN = "stock_con_precios.xlsx"
SALIDA = "plantilla_carga_stock.xlsx"
FILAS_EXTRA_PARA_PRODUCTOS_NUEVOS = 40

TALLAS_VALIDAS = ["8", "10", "12", "14", "XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "UNICA"]
HEADERS = ["nombre", "codigo", "categoria", "talla", "stock", "precio_venta", "precio_costo"]
COL = {h: i + 1 for i, h in enumerate(HEADERS)}  # 1-indexed, ej. COL["talla"] == 4

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
STOCK_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")  # amarillo suave
PRECIO_FALTANTE_FILL = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")  # naranja suave
THIN_BORDER = Border(*(Side(style="thin", color="D1D5DB"),) * 4)


def construir_hoja_instrucciones(wb):
    ws = wb.active
    ws.title = "Instrucciones"

    ws.column_dimensions["A"].width = 100
    ws.sheet_view.showGridLines = False

    titulo = ws.cell(row=1, column=1, value="TresQuartos — Planilla de conteo de stock")
    titulo.font = Font(bold=True, size=16, color="1F2937")

    pasos = [
        "",
        "¿Para qué es esta planilla?",
        "Para cargar el stock REAL de cada producto (el que hay hoy en el local), talle por talle.",
        "El objetivo es contar de verdad, no confirmar los números viejos — por eso la columna",
        "'stock' viene en blanco a propósito, aunque el resto de los datos ya estén precargados.",
        "",
        "Cómo completarla (hoja 'Stock'):",
        "1. Cada fila es UN talle de UN producto. Ya están cargados los productos que veníamos",
        "   vendiendo — solo hay que contar y escribir el número en la columna 'stock'.",
        "2. Si un producto/talle no tiene stock, escribir 0 (no dejar la celda vacía).",
        "3. Si encuentran un producto que NO está en la lista, agregarlo en una fila nueva al final",
        "   (hay filas en blanco preparadas para eso). Completar nombre, categoría, talla, stock y precio.",
        "4. La columna 'codigo' es el código que usa IMAGO (el proveedor) para ese producto — se ve",
        "   en el Cuadro de Pedido. Es OPCIONAL: completarlo solo si están seguros de que es el mismo",
        "   producto. Si no lo saben o no están seguros, dejarlo en blanco — no hay que adivinar.",
        "5. La columna 'categoria' tiene una lista para elegir (clic en la celda y aparece la flechita).",
        "   Si es una categoría nueva, se puede escribir directamente.",
        "6. La columna 'talla' también tiene una lista con los talles habituales. Si el talle no está",
        "   en la lista, escribirlo igual (ej. talles especiales).",
        "7. Las celdas de la columna 'precio_venta' en naranja son productos sin precio cargado:",
        "   completar el precio de venta actual. Las que no están en naranja ya tienen un precio",
        "   de referencia — revisar si sigue siendo correcto y corregirlo si cambió.",
        "8. La columna 'precio_costo' es opcional, se puede dejar en blanco.",
        "9. No hace falta ordenar ni acomodar nada — con que los números estén bien alcanza.",
        "",
        "¿Qué pasa después?",
        "Una vez completa, le devuelven este mismo archivo a Agustín. Él lo sube al sistema con",
        "una herramienta que ya está preparada: primero revisa las diferencias contra lo que hay",
        "cargado (sin tocar nada todavía) y recién después de confirmar, actualiza el stock real",
        "de la tienda. No hace falta que hagan nada más técnico que completar los números.",
        "",
        "Dudas: consultar a Agustín antes de inventar un formato distinto — con que sea prolijo",
        "y completo alcanza, no hace falta que quede 'lindo'.",
    ]

    for i, linea in enumerate(pasos, start=2):
        cell = ws.cell(row=i, column=1, value=linea)
        if linea.endswith(":") or linea.startswith("¿"):
            cell.font = Font(bold=True, size=12, color="1F2937")
        else:
            cell.font = Font(size=11, color="374151")
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def construir_hoja_stock(wb):
    df = pd.read_excel(ORIGEN)
    categorias = sorted(df["categoria"].dropna().unique().tolist())
    n_cols = len(HEADERS)

    ws = wb.create_sheet("Stock")
    for h, col in COL.items():
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(ord('A') + n_cols - 1)}1"

    fila = 2
    for _, row in df.iterrows():
        nombre = str(row["nombre"]).strip()
        categoria = str(row["categoria"]).strip()
        talla = str(row["talla"]).strip()
        precio_venta = row["precio_venta"] if pd.notna(row["precio_venta"]) and row["precio_venta"] != 0 else None

        ws.cell(row=fila, column=COL["nombre"], value=nombre)
        ws.cell(row=fila, column=COL["codigo"], value=None)  # sin código conocido: se completa a mano si corresponde
        ws.cell(row=fila, column=COL["categoria"], value=categoria)
        ws.cell(row=fila, column=COL["talla"], value=talla)
        stock_cell = ws.cell(row=fila, column=COL["stock"], value=None)
        stock_cell.fill = STOCK_FILL
        precio_cell = ws.cell(row=fila, column=COL["precio_venta"], value=precio_venta)
        if precio_venta is None:
            precio_cell.fill = PRECIO_FALTANTE_FILL
        ws.cell(row=fila, column=COL["precio_costo"], value=None)

        for col in range(1, n_cols + 1):
            ws.cell(row=fila, column=col).border = THIN_BORDER
        fila += 1

    ultima_fila_datos = fila - 1

    # Filas en blanco al final para productos nuevos
    for _ in range(FILAS_EXTRA_PARA_PRODUCTOS_NUEVOS):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=fila, column=col)
            cell.border = THIN_BORDER
            if col == COL["stock"]:
                cell.fill = STOCK_FILL
        fila += 1

    ultima_fila_total = fila - 1

    # Anchos de columna
    anchos = {"A": 34, "B": 12, "C": 24, "D": 10, "E": 10, "F": 14, "G": 14}
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    # Validación de datos: categoría (lista + libre) y talla (lista + libre).
    # La lista de categorías (21 valores) supera el límite de 255 caracteres
    # que Excel permite para una lista inline ("a,b,c"), así que se referencia
    # un rango en una hoja auxiliar oculta en vez de escribir la lista en la
    # fórmula de validación.
    aux = wb.create_sheet("_categorias")
    for i, cat in enumerate(categorias, start=1):
        aux.cell(row=i, column=1, value=cat)
    aux.sheet_state = "hidden"

    col_categoria = chr(ord("A") + COL["categoria"] - 1)
    col_talla = chr(ord("A") + COL["talla"] - 1)

    dv_categoria = DataValidation(
        type="list",
        formula1=f"_categorias!$A$1:$A${len(categorias)}",
        allow_blank=True,
    )
    dv_categoria.showErrorMessage = False  # permite escribir una categoría nueva
    ws.add_data_validation(dv_categoria)
    dv_categoria.add(f"{col_categoria}2:{col_categoria}{ultima_fila_total}")

    lista_tallas = ",".join(TALLAS_VALIDAS)
    dv_talla = DataValidation(type="list", formula1=f'"{lista_tallas}"', allow_blank=True)
    dv_talla.showErrorMessage = False  # permite escribir un talle no listado
    ws.add_data_validation(dv_talla)
    dv_talla.add(f"{col_talla}2:{col_talla}{ultima_fila_total}")

    print(f"Productos existentes: {ultima_fila_datos - 1} filas")
    print(f"Filas en blanco para productos nuevos: {FILAS_EXTRA_PARA_PRODUCTOS_NUEVOS}")
    print(f"Categorías detectadas: {len(categorias)}")
    print(f"Precios faltantes (celda naranja): {df['precio_venta'].isna().sum() + (df['precio_venta'] == 0).sum()}")


def main():
    wb = Workbook()
    construir_hoja_instrucciones(wb)
    construir_hoja_stock(wb)
    wb.save(SALIDA)
    print(f"\nListo: {SALIDA}")


if __name__ == "__main__":
    main()
